"""
Extraction layer for Approach G (Gemini Files API + Backend Entity Resolution).

PDFs   → raw bytes returned as-is; Gemini Files API handles native parsing.
        FOS/MR name detected cheaply from first two pages only.

Tabular (CSV / XLSX / XLS) → converted to minimal CSV text for direct LLM input.
        FOS name detected from meta-rows above the header.
"""
from __future__ import annotations

import csv
import io
import logging
import re
import time
from dataclasses import dataclass
from pathlib import Path

logger = logging.getLogger(__name__)


_FOS_KEYWORDS = re.compile(
    r"\b(fos|fos\s*name|mr|mr\s*name|rep|representative|field\s*officer|medical\s*rep)\b",
    re.IGNORECASE,
)


@dataclass
class ExtractionResult:
    raw_text: str | None          # tabular files — minimal CSV text
    raw_bytes: bytes | None       # PDF — sent as-is to Gemini Files API
    is_pdf: bool
    detected_fos_name: str | None = None
    total_pages: int | None = None
    total_rows: int | None = None


# ---------------------------------------------------------------------------
# PDF
# ---------------------------------------------------------------------------

def extract_pdf(
    content: bytes,
    *,
    detect_fos: bool = True,
    log_prefix: str = "",
) -> ExtractionResult:
    """
    PDF extraction. When `detect_fos` is False we skip opening the PDF entirely
    (downstream parsing handles binary directly), avoiding a redundant
    pdfplumber pass on top of `probe_size`.
    """
    if not detect_fos:
        logger.info(
            "%s extract_pdf: skipping local parse (detect_fos=False) bytes=%d",
            log_prefix, len(content),
        )
        return ExtractionResult(
            raw_text=None,
            raw_bytes=content,
            is_pdf=True,
            detected_fos_name=None,
            total_pages=None,
        )

    import pdfplumber

    t0 = time.perf_counter()
    fos_name: str | None = None
    with pdfplumber.open(io.BytesIO(content)) as pdf:
        total_pages = len(pdf.pages)
        head_lines: list[str] = []
        for page in pdf.pages[:2]:
            t = page.extract_text(x_tolerance=3, y_tolerance=3) or ""
            head_lines.extend(t.splitlines())
        fos_name = _detect_fos_in_lines(head_lines[:30])

    elapsed_ms = (time.perf_counter() - t0) * 1000
    logger.info(
        "%s extract_pdf: pages=%d fos_detected=%s elapsed_ms=%.0f",
        log_prefix, total_pages, bool(fos_name), elapsed_ms,
    )

    return ExtractionResult(
        raw_text=None,
        raw_bytes=content,
        is_pdf=True,
        detected_fos_name=fos_name,
        total_pages=total_pages,
    )


# ---------------------------------------------------------------------------
# CSV
# ---------------------------------------------------------------------------

def extract_csv(content: bytes, *, log_prefix: str = "") -> ExtractionResult:
    t0 = time.perf_counter()
    decoded = content.decode("utf-8", errors="replace")
    all_rows = [
        [c if c is not None else "" for c in row]
        for row in csv.reader(io.StringIO(decoded))
    ]
    res = _sheets_to_result([("csv", all_rows)])
    logger.info(
        "%s extract_csv: rows_in=%d rows_out=%s fos_detected=%s elapsed_ms=%.0f",
        log_prefix, len(all_rows), res.total_rows,
        bool(res.detected_fos_name), (time.perf_counter() - t0) * 1000,
    )
    return res


# ---------------------------------------------------------------------------
# XLSX  (reads ALL sheets; normal mode so a wrong <dimension> can't blind us)
# ---------------------------------------------------------------------------

def extract_xlsx(content: bytes, *, log_prefix: str = "") -> ExtractionResult:
    import openpyxl

    t0 = time.perf_counter()
    # read_only=True trusts the file's declared <dimension>. Some report
    # exporters (e.g. "A2E_Engine") write it incorrectly as "A1:B1", which
    # silently hid every row. Normal mode parses the actual cells.
    wb = openpyxl.load_workbook(io.BytesIO(content), read_only=False, data_only=True)
    sheets = _collect_sheets_openpyxl(wb)
    wb.close()
    res = _sheets_to_result(sheets)
    logger.info(
        "%s extract_xlsx: sheets_with_data=%d rows_out=%s fos_detected=%s elapsed_ms=%.0f",
        log_prefix, len(sheets), res.total_rows,
        bool(res.detected_fos_name), (time.perf_counter() - t0) * 1000,
    )
    return res


def _collect_sheets_openpyxl(wb) -> list[tuple[str, list[list[str]]]]:
    sheets: list[tuple[str, list[list[str]]]] = []
    for ws in wb.worksheets:
        rows = [
            [("" if c is None else str(c)) for c in row]
            for row in ws.iter_rows(values_only=True)
        ]
        if any(any(c.strip() for c in r) for r in rows):
            sheets.append((ws.title, rows))
    return sheets


# ---------------------------------------------------------------------------
# XLS  (legacy OLE2; reads ALL sheets)
# ---------------------------------------------------------------------------

def extract_xls(content: bytes, *, log_prefix: str = "") -> ExtractionResult:
    import xlrd

    t0 = time.perf_counter()
    wb = xlrd.open_workbook(file_contents=content)
    sheets: list[tuple[str, list[list[str]]]] = []
    for si in range(wb.nsheets):
        ws = wb.sheet_by_index(si)
        rows = [
            [_xls_cell_str(ws.cell_value(r, c)) for c in range(ws.ncols)]
            for r in range(ws.nrows)
        ]
        if any(any(c.strip() for c in r) for r in rows):
            sheets.append((ws.name, rows))
    res = _sheets_to_result(sheets)
    logger.info(
        "%s extract_xls: sheets_with_data=%d rows_out=%s fos_detected=%s elapsed_ms=%.0f",
        log_prefix, len(sheets), res.total_rows,
        bool(res.detected_fos_name), (time.perf_counter() - t0) * 1000,
    )
    return res


def _xls_cell_str(v: object) -> str:
    # xlrd returns every number as float; keep integers clean ("10" not "10.0").
    if isinstance(v, float) and v == int(v):
        return str(int(v))
    return "" if v is None else str(v)


# ---------------------------------------------------------------------------
# Tabular helper — combine one-or-more sheets into faithful CSV text
# ---------------------------------------------------------------------------

def _sheets_to_result(
    sheets: list[tuple[str, list[list[str]]]],
) -> ExtractionResult:
    """
    Render every sheet that has data into CSV text, faithfully: title rows,
    header, and data are all preserved. The LLM needs the title rows (they
    carry the report month for dateless files) and needs every sheet.

    Multiple sheets are separated by a "### SHEET: <name>" marker so the model
    can tell them apart.
    """
    out = io.StringIO()
    writer = csv.writer(out)
    multi = len(sheets) > 1
    total_rows = 0
    fos_name: str | None = None

    for name, rows in sheets:
        rows = list(rows)
        while rows and not any(c.strip() for c in rows[-1]):
            rows.pop()
        if not rows:
            continue
        if multi:
            out.write(f"### SHEET: {name}\n")
        header_idx = _find_header_row(rows)
        if fos_name is None:
            fos_name = _find_fos_in_rows(rows, header_idx)
        for row in rows:
            if any(c.strip() for c in row):
                writer.writerow(row)
                total_rows += 1

    return ExtractionResult(
        raw_text=out.getvalue(),
        raw_bytes=None,
        is_pdf=False,
        detected_fos_name=fos_name,
        total_rows=total_rows,
    )


# ---------------------------------------------------------------------------
# Shared helpers
# ---------------------------------------------------------------------------

def _find_header_row(rows: list[list[str]]) -> int:
    for i, row in enumerate(rows[:25]):
        string_cells = [c for c in row if c.strip() and not _is_numeric(c)]
        if len(string_cells) >= 3:
            return i
    return 0


def _is_numeric(value: str) -> bool:
    try:
        float(value.replace(",", ""))
        return True
    except ValueError:
        return False


def _find_fos_in_rows(rows: list[list[str]], header_row_index: int) -> str | None:
    for row in rows[:header_row_index]:
        for i, cell in enumerate(row):
            if cell and _FOS_KEYWORDS.search(cell):
                after_colon = re.split(r"[:：]", cell, maxsplit=1)
                if len(after_colon) == 2 and after_colon[1].strip():
                    return after_colon[1].strip()
                if i + 1 < len(row) and row[i + 1].strip():
                    return row[i + 1].strip()
    return None


def _detect_fos_in_lines(lines: list[str]) -> str | None:
    for line in lines:
        if _FOS_KEYWORDS.search(line):
            after_colon = re.split(r"[:：]", line, maxsplit=1)
            if len(after_colon) == 2 and after_colon[1].strip():
                return after_colon[1].strip()
    return None


# ---------------------------------------------------------------------------
# Dispatcher + size introspection
# ---------------------------------------------------------------------------

def _sniff_format(filename: str, content: bytes) -> str:
    """
    Determine the real format from content (magic bytes), falling back to the
    filename extension. Distributors mislabel files — e.g. a `.xls` that is
    actually an OOXML `.xlsx` — and trusting the extension alone crashes the
    wrong parser (xlrd on a zip). Returns one of:
    pdf | xlsx | xls | csv | numbers | ods | <raw-ext>.
    """
    head = content[:8]
    if head.startswith(b"%PDF"):
        return "pdf"
    if head.startswith(b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1"):
        return "xls"  # OLE2 compound document (true legacy .xls)
    if head.startswith(b"PK\x03\x04"):
        # Zip container: OOXML xlsx, ODS, or Apple .numbers.
        try:
            import zipfile
            names = zipfile.ZipFile(io.BytesIO(content)).namelist()
        except Exception:  # noqa: BLE001
            return "xlsx"
        if any(n.startswith("xl/") for n in names):
            return "xlsx"
        if any(n.endswith(".iwa") for n in names):
            return "numbers"
        if "content.xml" in names:
            return "ods"
        return "xlsx"
    return Path(filename).suffix.lower().lstrip(".") or "csv"


def extract(
    filename: str,
    content: bytes,
    *,
    detect_fos: bool = True,
    log_prefix: str = "",
) -> ExtractionResult:
    fmt = _sniff_format(filename, content)
    ext = Path(filename).suffix.lower().lstrip(".")
    logger.info(
        "%s extract: filename=%r ext=%s sniffed=%s bytes=%d detect_fos=%s",
        log_prefix, filename, ext, fmt, len(content), detect_fos,
    )
    if fmt != ext and fmt in ("pdf", "xlsx", "xls"):
        logger.warning(
            "%s extract: extension/content mismatch — file is %r despite .%s extension",
            log_prefix, fmt, ext,
        )
    if fmt == "pdf":
        return extract_pdf(content, detect_fos=detect_fos, log_prefix=log_prefix)
    if fmt == "csv":
        return extract_csv(content, log_prefix=log_prefix)
    if fmt == "xlsx":
        return extract_xlsx(content, log_prefix=log_prefix)
    if fmt == "xls":
        return extract_xls(content, log_prefix=log_prefix)
    if fmt in ("numbers", "ods"):
        raise ValueError(
            f"Unsupported file format: {fmt!r}. Export to .xlsx or .csv and re-upload."
        )
    raise ValueError(f"Unsupported file format: {ext!r}. Supported: pdf, csv, xlsx, xls")


def probe_size(filename: str, content: bytes) -> tuple[int | None, int | None]:
    """
    Cheap size probe used by the upload endpoint to enforce limits BEFORE heavy processing.
    Returns (page_count_or_None, row_count_or_None). Callers check len(content) separately.

    Uses content-sniffed format (not the extension) and, for spreadsheets, counts
    rows across ALL sheets in normal mode so a wrong <dimension> can't under-report.
    """
    fmt = _sniff_format(filename, content)
    t0 = time.perf_counter()
    pages: int | None = None
    rows: int | None = None
    try:
        if fmt == "pdf":
            import pdfplumber
            with pdfplumber.open(io.BytesIO(content)) as pdf:
                pages = len(pdf.pages)
        elif fmt == "csv":
            decoded = content.decode("utf-8", errors="replace")
            rows = sum(1 for _ in csv.reader(io.StringIO(decoded)))
        elif fmt == "xlsx":
            import openpyxl
            wb = openpyxl.load_workbook(io.BytesIO(content), read_only=False, data_only=True)
            rows = sum(ws.max_row or 0 for ws in wb.worksheets)
            wb.close()
        elif fmt == "xls":
            import xlrd
            wb = xlrd.open_workbook(file_contents=content)
            rows = sum(wb.sheet_by_index(i).nrows for i in range(wb.nsheets))
    except Exception as exc:  # noqa: BLE001
        logger.warning(
            "probe_size: failed for %r (%s: %s) — returning None,None",
            filename, type(exc).__name__, exc,
        )
        return None, None
    logger.info(
        "probe_size: filename=%r fmt=%s bytes=%d pages=%s rows=%s elapsed_ms=%.0f",
        filename, fmt, len(content), pages, rows, (time.perf_counter() - t0) * 1000,
    )
    return pages, rows
